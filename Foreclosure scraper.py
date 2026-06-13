import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import threading
import time
import random
import os
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    messagebox.showerror("Missing Library", "openpyxl not found.\nRun: pip install openpyxl")
    raise SystemExit

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options
    from selenium.common.exceptions import (
        TimeoutException, NoSuchElementException, WebDriverException
    )
except ImportError:
    messagebox.showerror("Missing Library", "selenium not found.\nRun: pip install selenium")
    raise SystemExit

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


XPATHS = {
    "Case":      "//p[contains(text(),'Case #:')]/following-sibling::p",
    "Sale Date": "//p[contains(text(),'Sale Date:')]/following-sibling::p",
    "County":    "//p[contains(text(),'County:')]/following-sibling::p",
    "State":     "//p[contains(text(),'State:')]/following-sibling::p",
    "Address":   "//p[contains(text(),'Address:')]/following-sibling::p",
    "Book Page": "//p[contains(text(),'Book Page:')]/following-sibling::p",
    "Court SP":  "//p[contains(text(),'Court SP #:')]/following-sibling::p",
}

COLLECT_ORDER     = ["Case", "Sale Date", "County", "State", "Address", "Book Page", "Court SP"]
EXCEL_HEADERS     = ["County", "Sale Date", "State", "Court SP", "Case", "Address", "Book Page"]
PAGINATION_XPATH  = '(//a[contains(text(),"Next ")])[2]'
PAGE_LOAD_TIMEOUT = 30
RANDOM_WAIT_MIN   = 5
RANDOM_WAIT_MAX   = 10

COLORS = {
    "bg_dark":        "#0D1117",
    "bg_panel":       "#161B22",
    "bg_card":        "#1C2128",
    "accent":         "#58A6FF",
    "accent2":        "#3FB950",
    "warning":        "#D29922",
    "danger":         "#F85149",
    "pause_col":      "#E3B341",
    "text_primary":   "#E6EDF3",
    "text_secondary": "#8B949E",
    "border":         "#30363D",
    "header_bg":      "#1F2937",
}


class ExcelWriter:
    def __init__(self, filepath):
        self.filepath = filepath
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Foreclosure Data"
        self._write_headers()
        self.row  = 2
        self._lock = threading.Lock()

    def _write_headers(self):
        header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill  = PatternFill("solid", fgColor="1F3A5F")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin   = Side(style="thin", color="2E4A6B")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        col_widths = {
            "County": 14, "Sale Date": 14, "State": 8,
            "Court SP": 20, "Case": 20, "Address": 35, "Book Page": 14
        }
        for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
            cell = self.ws.cell(row=1, column=col_idx, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = border
            self.ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = col_widths.get(header, 16)
        self.ws.row_dimensions[1].height = 28
        self.ws.freeze_panes = "A2"

    def write_row(self, data):
        thin   = Side(style="thin", color="3D4F61")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        even_fill = PatternFill("solid", fgColor="F0F4F8") if self.row % 2 == 0 else None
        with self._lock:
            for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
                cell = self.ws.cell(row=self.row, column=col_idx, value=data.get(header, ""))
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.border    = border
                if even_fill:
                    cell.fill = even_fill
            self.ws.row_dimensions[self.row].height = 20
            self.row += 1
            self._save()

    def _save(self):
        try:
            self.wb.save(self.filepath)
        except PermissionError:
            pass

    def finalize(self):
        with self._lock:
            self._save()


class ScraperEngine:
    def __init__(self, urls, output_path, callbacks):
        self.urls        = urls
        self.output_path = output_path
        self.on_log      = callbacks.get("log",    lambda m, t="info": None)
        self.on_count    = callbacks.get("count",  lambda n: None)
        self.on_status   = callbacks.get("status", lambda s: None)
        self.on_done     = callbacks.get("done",   lambda: None)
        self.driver      = None
        self.excel       = None
        self._paused     = threading.Event()
        self._stopped    = threading.Event()
        self._paused.set()
        self.total_count = 0

    def pause(self):
        self._paused.clear()
        self.on_log("Scraper paused.", "warning")
        self.on_status("PAUSED")

    def resume(self):
        self._paused.set()
        self.on_log("Scraper resumed.", "success")
        self.on_status("RUNNING")

    def stop(self):
        self._stopped.set()
        self._paused.set()
        self.on_log("Stop signal received. Finalizing...", "danger")

    def _init_browser(self):
        opts = Options()
        opts.add_argument("--start-maximized")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_experimental_option("useAutomationExtension", False)
        self.driver = webdriver.Chrome(options=opts)
        self.on_log("Browser started.", "info")

    def _close_browser(self):
        try:
            if self.driver:
                self.driver.quit()
                self.driver = None
        except Exception:
            pass

    def _wait_for_page(self):
        WebDriverWait(self.driver, PAGE_LOAD_TIMEOUT).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )

    def _check_pause_stop(self):
        self._paused.wait()
        return self._stopped.is_set()

    def _collect_page_data(self):
        raw = {}
        for field in COLLECT_ORDER:
            try:
                elements = self.driver.find_elements(By.XPATH, XPATHS[field])
                raw[field] = [el.text.strip() for el in elements if el.text.strip()]
            except Exception:
                raw[field] = []

        max_len = max((len(v) for v in raw.values()), default=0)
        if max_len == 0:
            return []

        records = []
        for i in range(max_len):
            record = {}
            for header in EXCEL_HEADERS:
                vals = raw.get(header, [])
                record[header] = vals[i] if i < len(vals) else ""
            records.append(record)
        return records

    def _click_next(self):
        try:
            btn = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, PAGINATION_XPATH))
            )
            self.driver.execute_script("arguments[0].click();", btn)
            return True
        except (TimeoutException, NoSuchElementException):
            return False

    def run(self):
        try:
            self._init_browser()
            self.excel = ExcelWriter(self.output_path)
            self.on_status("RUNNING")

            for url_idx, url in enumerate(self.urls, start=1):
                if self._stopped.is_set():
                    break
                url = url.strip()
                if not url:
                    continue

                self.on_log(f"[{url_idx}/{len(self.urls)}] Opening: {url}", "info")

                try:
                    self.driver.get(url)
                    self._wait_for_page()
                except WebDriverException as e:
                    self.on_log(f"Failed to load URL: {e}", "warning")
                    continue

                page_num = 1

                while True:
                    if self._check_pause_stop():
                        break

                    self.on_log(f"Page {page_num} - collecting data...", "info")
                    records = self._collect_page_data()

                    for rec in records:
                        if self._check_pause_stop():
                            break
                        self.excel.write_row(rec)
                        self.total_count += 1
                        self.on_count(self.total_count)

                    self.on_log(f"Page {page_num}: {len(records)} record(s) collected.", "success")

                    if self._stopped.is_set():
                        break

                    if not self._click_next():
                        self.on_log("No more pages for this URL.", "info")
                        break

                    try:
                        self._wait_for_page()
                    except TimeoutException:
                        self.on_log("Timeout waiting for next page.", "warning")
                        break

                    wait_secs = random.uniform(RANDOM_WAIT_MIN, RANDOM_WAIT_MAX)
                    self.on_log(f"Waiting {wait_secs:.1f}s before next page...", "info")
                    for _ in range(int(wait_secs * 10)):
                        if self._stopped.is_set():
                            break
                        time.sleep(0.1)

                    page_num += 1

        except Exception as e:
            self.on_log(f"Unexpected error: {e}", "danger")

        finally:
            if self.excel:
                self.excel.finalize()
                self.on_log(
                    f"Data saved -> {self.output_path}  ({self.total_count} records)", "success"
                )
            self._close_browser()
            self.on_status("IDLE")
            self.on_done()


class ForeClosureScraperApp:
    def __init__(self, root):
        self.root           = root
        self.engine         = None
        self.thread         = None
        self._output_path   = ""
        self._total_records = 0
        self._start_time    = None

        self._setup_window()
        self._build_ui()

    def _setup_window(self):
        self.root.title("Brock & Scott - Foreclosure Scraper")
        self.root.geometry("920x760")
        self.root.minsize(800, 660)
        self.root.configure(bg=COLORS["bg_dark"])
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth()  - 920) // 2
        y = (self.root.winfo_screenheight() - 760) // 2
        self.root.geometry(f"920x760+{x}+{y}")
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    def _build_ui(self):
        C = COLORS

        header = tk.Frame(self.root, bg="#0A0E14", height=56)
        header.pack(fill="x")
        header.pack_propagate(False)

        tk.Label(
            header,
            text="BROCK & SCOTT  -  FORECLOSURE DATA SCRAPER",
            font=("Courier New", 13, "bold"),
            bg="#0A0E14", fg=C["accent"],
        ).pack(side="left", padx=20, pady=16)

        self.status_var = tk.StringVar(value="IDLE")
        self.status_lbl = tk.Label(
            header,
            textvariable=self.status_var,
            font=("Courier New", 9, "bold"),
            bg="#1C2128", fg=C["text_secondary"],
            padx=10, pady=4,
            relief="flat",
        )
        self.status_lbl.pack(side="right", padx=20, pady=14)

        body = tk.Frame(self.root, bg=C["bg_dark"])
        body.pack(fill="both", expand=True, padx=16, pady=12)

        left  = tk.Frame(body, bg=C["bg_dark"], width=430)
        left.pack(side="left", fill="both", expand=True, padx=(0, 8))

        right = tk.Frame(body, bg=C["bg_dark"], width=440)
        right.pack(side="right", fill="both", expand=True, padx=(8, 0))

        self._card(left,  "INPUT URLS",      self._build_url_section)
        self._card(left,  "OUTPUT FILE",     self._build_output_section)
        self._card(left,  "CONTROLS",        self._build_controls)
        self._card(right, "LIVE STATISTICS", self._build_stats)
        self._card(right, "ACTIVITY LOG",    self._build_log, expand=True)

    def _card(self, parent, title, builder_fn, expand=False):
        C = COLORS
        frame = tk.Frame(
            parent,
            bg=C["bg_panel"],
            highlightbackground=C["border"],
            highlightthickness=1,
        )
        pack_kw = {"fill": "both", "pady": (0, 10)}
        if expand:
            pack_kw["expand"] = True
        frame.pack(**pack_kw)

        hdr = tk.Frame(frame, bg=C["header_bg"], height=32)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(
            hdr,
            text=title,
            font=("Courier New", 9, "bold"),
            bg=C["header_bg"], fg=C["accent"],
            padx=12,
        ).pack(side="left", pady=7)

        body = tk.Frame(frame, bg=C["bg_panel"])
        body.pack(fill="both", expand=True, padx=12, pady=10)
        builder_fn(body)

    def _build_url_section(self, parent):
        C = COLORS
        tk.Label(
            parent,
            text="Paste URLs (one per line):",
            font=("Courier New", 8),
            bg=C["bg_panel"], fg=C["text_secondary"],
        ).pack(anchor="w")

        self.url_text = tk.Text(
            parent,
            height=6,
            bg=C["bg_card"], fg=C["text_primary"],
            insertbackground=C["accent"],
            font=("Courier New", 9),
            relief="flat",
            highlightbackground=C["border"],
            highlightthickness=1,
            wrap="none",
        )
        self.url_text.pack(fill="x", pady=(4, 8))

        tk.Label(
            parent,
            text="- OR upload an Excel file with URLs -",
            font=("Courier New", 8),
            bg=C["bg_panel"], fg=C["text_secondary"],
        ).pack()

        btn_row = tk.Frame(parent, bg=C["bg_panel"])
        btn_row.pack(fill="x", pady=(6, 0))

        self._btn(btn_row, "Upload Excel", self._upload_excel, C["accent"], side="left")

        self.excel_lbl = tk.Label(
            btn_row,
            text="No file selected",
            font=("Courier New", 8),
            bg=C["bg_panel"], fg=C["text_secondary"],
        )
        self.excel_lbl.pack(side="left", padx=10)

    def _upload_excel(self):
        path = filedialog.askopenfilename(
            title="Select Excel with URLs",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("CSV", "*.csv"), ("All", "*.*")]
        )
        if not path:
            return
        try:
            if PANDAS_AVAILABLE:
                df = pd.read_excel(path) if path.endswith((".xlsx", ".xls")) else pd.read_csv(path)
                urls = []
                for col in df.columns:
                    if any(k in col.lower() for k in ("url", "link", "http")):
                        urls = df[col].dropna().astype(str).tolist()
                        break
                if not urls:
                    urls = df.iloc[:, 0].dropna().astype(str).tolist()
            else:
                wb   = openpyxl.load_workbook(path)
                ws   = wb.active
                urls = [
                    str(row[0].value).strip()
                    for row in ws.iter_rows(min_row=2)
                    if row[0].value
                ]
            self.url_text.delete("1.0", "end")
            self.url_text.insert("1.0", "\n".join(urls))
            self.excel_lbl.config(
                text=f"Loaded: {os.path.basename(path)} ({len(urls)} URLs)",
                fg=COLORS["accent2"],
            )
            self._log(f"Loaded {len(urls)} URL(s) from Excel.", "success")
        except Exception as e:
            messagebox.showerror("Error", f"Could not read file:\n{e}")

    def _build_output_section(self, parent):
        C = COLORS
        row = tk.Frame(parent, bg=C["bg_panel"])
        row.pack(fill="x")
        self.output_var = tk.StringVar(value=self._default_output_path())
        entry = tk.Entry(
            row,
            textvariable=self.output_var,
            font=("Courier New", 8),
            bg=C["bg_card"], fg=C["text_primary"],
            insertbackground=C["accent"],
            relief="flat",
            highlightbackground=C["border"],
            highlightthickness=1,
        )
        entry.pack(side="left", fill="x", expand=True, ipady=5, padx=(0, 6))
        self._btn(row, "Browse", self._choose_output, C["border"], side="left", padx=0)

    def _choose_output(self):
        path = filedialog.asksaveasfilename(
            title="Save Excel As",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"foreclosure_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
        if path:
            self.output_var.set(path)

    def _default_output_path(self):
        desktop = Path.home() / "Desktop"
        folder  = desktop if desktop.exists() else Path.home()
        return str(folder / f"foreclosure_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    def _build_controls(self, parent):
        C = COLORS
        row = tk.Frame(parent, bg=C["bg_panel"])
        row.pack()
        self.btn_start  = self._btn(row, "START",  self._start,  C["accent2"],   side="left", padx=4)
        self.btn_pause  = self._btn(row, "PAUSE",  self._pause,  C["pause_col"], side="left", padx=4, state="disabled")
        self.btn_resume = self._btn(row, "RESUME", self._resume, C["accent"],    side="left", padx=4, state="disabled")
        self.btn_stop   = self._btn(row, "STOP",   self._stop,   C["danger"],    side="left", padx=4, state="disabled")

    def _build_stats(self, parent):
        C = COLORS
        self.count_var = tk.StringVar(value="0")
        tk.Label(
            parent,
            textvariable=self.count_var,
            font=("Courier New", 40, "bold"),
            bg=C["bg_panel"], fg=C["accent2"],
        ).pack()
        tk.Label(
            parent,
            text="RECORDS COLLECTED",
            font=("Courier New", 8, "bold"),
            bg=C["bg_panel"], fg=C["text_secondary"],
        ).pack()
        tk.Frame(parent, bg=C["border"], height=1).pack(fill="x", pady=10)
        self.elapsed_var = tk.StringVar(value="Elapsed: --")
        tk.Label(
            parent,
            textvariable=self.elapsed_var,
            font=("Courier New", 8),
            bg=C["bg_panel"], fg=C["text_secondary"],
        ).pack()

    def _build_log(self, parent):
        C = COLORS
        self.log_box = scrolledtext.ScrolledText(
            parent,
            bg=C["bg_card"], fg=C["text_primary"],
            font=("Courier New", 8),
            relief="flat",
            state="disabled",
            wrap="word",
            highlightbackground=C["border"],
            highlightthickness=1,
        )
        self.log_box.pack(fill="both", expand=True)
        self.log_box.tag_config("info",    foreground=C["text_secondary"])
        self.log_box.tag_config("success", foreground=C["accent2"])
        self.log_box.tag_config("warning", foreground=C["warning"])
        self.log_box.tag_config("danger",  foreground=C["danger"])
        self._log("Ready. Paste URLs and press START.", "info")

    def _log(self, msg, tag="info"):
        def _do():
            ts = datetime.now().strftime("%H:%M:%S")
            self.log_box.config(state="normal")
            self.log_box.insert("end", f"[{ts}] {msg}\n", tag)
            self.log_box.see("end")
            self.log_box.config(state="disabled")
        self.root.after(0, _do)

    def _btn(self, parent, text, cmd, color, side="left", padx=4, state="normal"):
        b = tk.Button(
            parent,
            text=text,
            command=cmd,
            state=state,
            font=("Courier New", 8, "bold"),
            bg=color,
            fg="#0D1117",
            activebackground=color,
            activeforeground="#0D1117",
            relief="flat",
            cursor="hand2",
            padx=10, pady=6,
        )
        b.pack(side=side, padx=padx)
        return b

    def _start(self):
        urls_raw = self.url_text.get("1.0", "end").strip()
        urls = [u.strip() for u in urls_raw.splitlines() if u.strip()]
        if not urls:
            messagebox.showwarning("No URLs", "Please paste at least one URL.")
            return
        output = self.output_var.get().strip()
        if not output:
            messagebox.showwarning("No Output", "Please specify an output file path.")
            return
        os.makedirs(os.path.dirname(output) or ".", exist_ok=True)
        self._total_records = 0
        self.count_var.set("0")
        self._start_time = time.time()
        self._tick_elapsed()
        self._set_buttons(running=True, paused=False)
        self._update_status("RUNNING", COLORS["accent2"])
        self.engine = ScraperEngine(
            urls=urls,
            output_path=output,
            callbacks={
                "log":    self._log,
                "count":  self._on_count,
                "status": self._on_status_cb,
                "done":   self._on_done,
            }
        )
        self.thread = threading.Thread(target=self.engine.run, daemon=True)
        self.thread.start()
        self._log(f"Started scraping {len(urls)} URL(s).", "success")

    def _pause(self):
        if self.engine:
            self.engine.pause()
            self._set_buttons(running=True, paused=True)
            self._update_status("PAUSED", COLORS["pause_col"])

    def _resume(self):
        if self.engine:
            self.engine.resume()
            self._set_buttons(running=True, paused=False)
            self._update_status("RUNNING", COLORS["accent2"])

    def _stop(self):
        if self.engine:
            self.engine.stop()
            self._set_buttons(running=False, paused=False)

    def _on_count(self, n):
        self.root.after(0, lambda: self.count_var.set(str(n)))

    def _on_status_cb(self, status):
        color_map = {
            "RUNNING": COLORS["accent2"],
            "PAUSED":  COLORS["pause_col"],
            "IDLE":    COLORS["text_secondary"],
        }
        self.root.after(
            0, lambda: self._update_status(status, color_map.get(status, COLORS["text_secondary"]))
        )

    def _on_done(self):
        self.root.after(0, self._finish)

    def _finish(self):
        self._set_buttons(running=False, paused=False)
        self._update_status("DONE", COLORS["accent"])
        output = self.output_var.get()
        self._log(f"All done! {self.engine.total_count} records saved to: {output}", "success")
        messagebox.showinfo(
            "Scraping Complete",
            f"{self.engine.total_count} records collected.\n\nSaved to:\n{output}"
        )

    def _set_buttons(self, running, paused):
        def _do():
            self.btn_start.config( state="disabled" if running else "normal")
            self.btn_pause.config( state="normal"   if (running and not paused) else "disabled")
            self.btn_resume.config(state="normal"   if paused else "disabled")
            self.btn_stop.config(  state="normal"   if running else "disabled")
        self.root.after(0, _do)

    def _update_status(self, text, color):
        self.status_var.set(text)
        self.status_lbl.config(fg=color)

    def _tick_elapsed(self):
        if self._start_time is None:
            return
        elapsed = int(time.time() - self._start_time)
        h, r = divmod(elapsed, 3600)
        m, s = divmod(r, 60)
        self.elapsed_var.set(f"Elapsed: {h:02d}:{m:02d}:{s:02d}")
        if self.thread and self.thread.is_alive():
            self.root.after(1000, self._tick_elapsed)

    def _on_close(self):
        if self.thread and self.thread.is_alive():
            if messagebox.askyesno("Quit?", "Scraper is running.\nStop it and exit?"):
                self._stop()
                self.root.after(1500, self.root.destroy)
        else:
            self.root.destroy()


if __name__ == "__main__":
    root = tk.Tk()
    app  = ForeClosureScraperApp(root)
    root.mainloop()
